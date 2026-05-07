#!/usr/bin/env python3
"""
identify_ej_aktiv.py — Identifierar kursstatus i tre lägen:

    - aktiv:   syns i aktuellt kursutbud
    - vilande: saknas i aktuellt utbud men kursplan finns fortfarande på du.se
    - ej-aktiv: varken i aktuellt utbud eller med publicerad kursplan på du.se

Arbetsflöde:
    1. Hämtar nuvarande ämnen + kurslistor från du.se (samma upptäcktslogik
         som scrape_hda_kursplaner.py — utan per-kurs-skrap).
    2. Jämför mot kursplansfiler i vault-dalarna-university/0X {INST}/Kursplaner/<AMNE>/.
    3. För kurser som inte går just nu verifieras kursplansidan:
             - kursplansida finns  -> taggas `vilande`
             - kursplansida saknas -> taggas `ej-aktiv`
    4. Re-emerged kurser taggas tillbaka till aktiv (tar bort status-taggar).
    5. Skapar/uppdaterar "Ej Aktiv <Ämne> MOC.md" per ämne.
    6. Skapar/uppdaterar 05 Analys/Vilande kursplaner.md + .xlsx.

Användning:
        python3 qa/identify_ej_aktiv.py                   # dry-run
        python3 qa/identify_ej_aktiv.py --apply           # skriv ändringar
        python3 qa/identify_ej_aktiv.py -i IIT            # bara IIT
        python3 qa/identify_ej_aktiv.py -s DTA            # bara Datateknik

Beroenden: requests, beautifulsoup4, openpyxl.
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from scrape_hda_kursplaner import (  # noqa: E402
    INSTITUTIONS,
    REQUEST_DELAY,
    SV_URL,
    discover_subjects_for_institution,
    discover_courses_for_subject,
    fetch_page,
)

VAULT = ROOT / "vault-dalarna-university"
INST_DIR_NAME = {"IIT": "01 IIT", "IHV": "02 IHV", "IKS": "03 IKS", "ISLL": "04 ISLL"}
VAULT_ANALYS = VAULT / "05 Analys"
VILANDE_ANALYS_MD = VAULT_ANALYS / "Vilande kursplaner.md"
VILANDE_ANALYS_XLSX = VAULT_ANALYS / "Vilande kursplaner.xlsx"

STATUS_EJ_AKTIV = "ej-aktiv"
STATUS_VILANDE = "vilande"
ALL_STATUS_TAGS = {STATUS_EJ_AKTIV, STATUS_VILANDE}


def kursplaner_dir(inst_code: str) -> Path:
    return VAULT / INST_DIR_NAME[inst_code] / "Kursplaner"

BOLD   = "\033[1m"
GREEN  = "\033[0;32m"
YELLOW = "\033[0;33m"
RED    = "\033[0;31m"
CYAN   = "\033[0;36m"
RESET  = "\033[0m"

COURSE_CODE_RE = re.compile(r"^[A-ZÅÄÖ0-9]{4,8}$")
KURSPLAN_URL = "https://www.du.se/sv/utbildning/kurser/kursplan/?code={code}"

HEADER_FILL = PatternFill("solid", fgColor="8B1A1A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="0563C1", underline="single")

DOWNLOAD_ICON_SVG = (
    '<svg class="download-xlsx-icon" width="16" height="16" viewBox="0 0 24 24" '
    'fill="none" stroke="currentColor" stroke-width="2" '
    'stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">'
    '<path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>'
    '<polyline points="7 10 12 15 17 10"/>'
    '<line x1="12" y1="15" x2="12" y2="3"/>'
    '</svg>'
)


# ─────────────────────────────────────────────────────────────────────────────
# Frontmatter helpers
# ─────────────────────────────────────────────────────────────────────────────

FM_RE = re.compile(r"^---\n(.*?)\n---\n", re.DOTALL)


def parse_frontmatter(text: str) -> tuple[dict, str, str] | None:
    """Returnerar (fm_dict, fm_block, body) eller None."""
    m = FM_RE.match(text)
    if not m:
        return None
    fm_block = m.group(1)
    body = text[m.end():]
    fm = {}
    for line in fm_block.split("\n"):
        if ":" in line and not line.startswith(" "):
            key, _, val = line.partition(":")
            fm[key.strip()] = val.strip()
    return fm, fm_block, body


def parse_list_field(value: str) -> list[str]:
    """Parsar `[a, b, c]` eller `a` → list[str]."""
    value = value.strip()
    if value.startswith("[") and value.endswith("]"):
        inner = value[1:-1]
        return [p.strip() for p in inner.split(",") if p.strip()]
    if value:
        return [value]
    return []


def serialize_list_field(items: list[str]) -> str:
    return "[" + ", ".join(items) + "]"


def update_fm_field(fm_block: str, key: str, new_value: str | None) -> str:
    """Skriver/lägger till/raderar en frontmatter-rad. new_value=None → radera."""
    pattern = re.compile(rf"^{re.escape(key)}:.*$", re.MULTILINE)
    if new_value is None:
        return pattern.sub("", fm_block).rstrip("\n") + "\n"
    new_line = f"{key}: {new_value}"
    if pattern.search(fm_block):
        return pattern.sub(new_line, fm_block)
    # append at end
    return fm_block.rstrip("\n") + "\n" + new_line


# ─────────────────────────────────────────────────────────────────────────────
# Vault-skanning
# ─────────────────────────────────────────────────────────────────────────────

def list_vault_codes_per_subject() -> dict[str, dict[str, Path]]:
    """Returnerar {subj_code -> {course_code -> path}} från vault-katalogerna,
    aggregerat över alla institutioner."""
    result: dict[str, dict[str, Path]] = defaultdict(dict)
    for inst_code in INST_DIR_NAME:
        kp = kursplaner_dir(inst_code)
        if not kp.exists():
            continue
        for subj_dir in kp.iterdir():
            if not subj_dir.is_dir():
                continue
            subj_code = subj_dir.name
            for f in subj_dir.glob("*.md"):
                if "MOC" in f.name:
                    continue
                stem = f.stem
                if COURSE_CODE_RE.match(stem):
                    result[subj_code][stem] = f
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Tagga / avtagga
# ─────────────────────────────────────────────────────────────────────────────

def set_course_status(path: Path, subject_name: str, status: str, apply: bool) -> bool:
    """Sätter kursstatus i frontmatter: aktiv/vilande/ej-aktiv.

    Returnerar True om filen ändrades.
    """
    if status not in {"active", STATUS_VILANDE, STATUS_EJ_AKTIV}:
        raise ValueError(f"Okänd status: {status}")

    text = path.read_text(encoding="utf-8")
    parsed = parse_frontmatter(text)
    if not parsed:
        return False
    _, fm_block, body = parsed

    new_fm = fm_block
    changed = False

    # tags
    tags_m = re.search(r"^tags:\s*(.+)$", new_fm, re.MULTILINE)
    tags = parse_list_field(tags_m.group(1)) if tags_m else []
    filtered_tags = [t for t in tags if t not in ALL_STATUS_TAGS]
    if status in ALL_STATUS_TAGS:
        filtered_tags.append(status)
    if filtered_tags != tags or not tags_m:
        new_fm = update_fm_field(new_fm, "tags", serialize_list_field(filtered_tags))
        changed = True

    # cssclasses
    css_m = re.search(r"^cssclasses:\s*(.+)$", new_fm, re.MULTILINE)
    css = parse_list_field(css_m.group(1)) if css_m else []
    filtered_css = [c for c in css if c not in ALL_STATUS_TAGS]
    if status in ALL_STATUS_TAGS:
        filtered_css.append(status)
    if filtered_css != css or (status in ALL_STATUS_TAGS and not css_m):
        if filtered_css:
            new_fm = update_fm_field(new_fm, "cssclasses", serialize_list_field(filtered_css))
        elif css_m:
            new_fm = update_fm_field(new_fm, "cssclasses", None)
        changed = True

    # up: ej-aktiv går till egen MOC, vilande/aktiv går till ämnes-MOC
    up_m = re.search(r'^up:\s*(.+)$', new_fm, re.MULTILINE)
    current_up = up_m.group(1).strip() if up_m else None
    target_up = (
        f'"[[Ej Aktiv {subject_name} MOC]]"'
        if status == STATUS_EJ_AKTIV
        else f'"[[{subject_name} MOC]]"'
    )
    if current_up != target_up:
        new_fm = update_fm_field(new_fm, "up", target_up)
        changed = True

    new_text = "---\n" + new_fm.rstrip("\n") + "\n---\n" + body
    if not changed or new_text == text:
        return False
    if apply:
        path.write_text(new_text, encoding="utf-8")
    return True


def read_course_name(path: Path) -> str:
    """Hämtar kursnamn från frontmatter (om möjligt)."""
    try:
        text = path.read_text(encoding="utf-8")
    except Exception:
        return ""
    for line in text.split("\n"):
        if line.startswith("kursnamn:"):
            return line.split(":", 1)[1].strip().strip('"')
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Ej-aktiv MOC
# ─────────────────────────────────────────────────────────────────────────────

def build_ej_aktiv_moc(subject_name: str, subject_code: str, inst_code: str,
                       orphan_paths: list[Path]) -> str:
    lines = [
        "---",
        "cssclasses: [ej-aktiv]",
        f"aliases: [Ej Aktiv {subject_name}]",
        f"tags: [MOC, ej-aktiv, {subject_code}, {inst_code}]",
        f'up: "[[{subject_name} MOC]]"',
        "---",
        "",
        f"# Ej Aktiv — {subject_name}",
        "",
        f"> Kurser inom {subject_name} som inte längre syns i du.se:s kursutbud.",
        "> Möjligen avvecklade, vilande eller omklassificerade.",
        "",
        f"## Kurser ({len(orphan_paths)} st)",
        "",
    ]
    for p in sorted(orphan_paths, key=lambda x: x.stem):
        course_name = read_course_name(p)
        if course_name:
            lines.append(f"- [[{p.stem}]] — {course_name}")
        else:
            lines.append(f"- [[{p.stem}]]")
    lines.append("")
    return "\n".join(lines)


def build_vilande_analysis_callout(rows: list[tuple[str, str, str, str, str]]) -> list[str]:
    """Bygg download-knapp + callout för Vilande kursplaner."""
    n = len(rows)
    href = VILANDE_ANALYS_XLSX.name.replace(" ", "-")
    lines = [
        f'<a class="download-xlsx" href="{href}" download>'
        f'{DOWNLOAD_ICON_SVG}'
        f'<span>Ladda ner som Excel-fil ({n} rader)</span>'
        f'</a>',
        "",
        f"> [!example]- {n} fynd — klicka för att expandera",
        ">",
        "> | Kursplan | Ämne | Institution | Problem |",
        "> | --- | --- | --- | --- |",
    ]
    for code, subj_code, inst_code, _subject_name, _course_name in rows:
        lines.append(
            "> | "
            f"[{code}]({KURSPLAN_URL.format(code=code)}) | "
            f"{subj_code} | {inst_code} | "
            "Kursplan finns på du.se men kursen saknas i aktuellt utbud |"
        )
    return lines


def replace_first_example_callout(text: str, new_block_lines: list[str]) -> str | None:
    """Ersätt första [!example]-blocket (inkl. ev. .xlsx-länk före blocket)."""
    lines = text.splitlines()
    start = None
    for i, line in enumerate(lines):
        if line.startswith("> [!example]"):
            start = i
            break
    if start is None:
        return None

    end = start + 1
    while end < len(lines) and lines[end].startswith(">"):
        end += 1

    block_start = start
    j = start - 1
    while j >= 0 and lines[j].strip() == "":
        j -= 1
    if j >= 0 and ".xlsx" in lines[j]:
        block_start = j

    merged = lines[:block_start] + new_block_lines + lines[end:]
    return "\n".join(merged) + ("\n" if text.endswith("\n") else "")


def build_vilande_analysis_template(callout_lines: list[str]) -> str:
    """Skapa grundmall för 05 Analys/Vilande kursplaner.md."""
    lines = [
        "---",
        "tags: [analys, kurslivscykel, vilande]",
        "up: \"[[Analys MOC]]\"",
        "status: första pass",
        "---",
        "",
        "# Vilande kursplaner",
        "",
        "## Problematiska kursplaner",
        "",
        *callout_lines,
        "",
        "## Syfte",
        "",
        "Identifiera kurser som är **vilande**: de erbjuds inte i aktuellt kursutbud men har fortfarande en publicerad kursplan på du.se. Det är ofta legitimt, men signalen kan också tyda på att utfasning eller arkivering inte genomförts klart.",
        "",
        "## Metod",
        "",
        "1. Läs in aktuellt kursutbud per ämne från du.se (sök/listning).",
        "2. Jämför mot kurskoder som finns i vaulten.",
        "3. För koder som saknas i aktuellt utbud: kontrollera direkt kursplans-URL.",
        "4. Markera som **vilande** när kursplanen fortfarande finns publicerad.",
        "",
        "**Begränsningar:**",
        "",
        "- En vilande kurs är inte automatiskt ett fel; analysen visar främst uppföljningsbehov.",
        "- Om du.se tillfälligt svarar fel kan klassningen bli osäker tills nästa körning.",
        "",
        "## Datakälla",
        "",
        "- `qa/identify_ej_aktiv.py`",
        "- Kursutbudslistning + kursplanssidor på du.se",
        "- Kursplansfiler under `0X {INST}/Kursplaner/`",
        "",
        "## Resultat",
        "",
        "*Fylls i efter genomgång.*",
        "",
        "## Observationer",
        "",
        "*Fylls i efter genomgång.*",
        "",
        "## Rekommendationer",
        "",
        "1. Bekräfta med ämnesföreträdare om varje vilande kurs ska återaktiveras, kvarstå eller avvecklas.",
        "2. För kurser som ska avvecklas: planera flytt till ej-aktiv eller annan tydlig arkiveringsstatus.",
        "3. Kör analysen regelbundet för att upptäcka glidning mellan utbud och publicerade kursplaner.",
        "",
    ]
    return "\n".join(lines)


def write_vilande_analysis(rows: list[tuple[str, str, str, str, str]], apply: bool) -> tuple[bool, bool]:
    """Skriv/uppdatera Vilande kursplaner.md + .xlsx. Returnerar (md_changed, xlsx_changed)."""
    rows = sorted(rows, key=lambda r: (r[2], r[1], r[0]))
    callout_lines = build_vilande_analysis_callout(rows)

    if VILANDE_ANALYS_MD.exists():
        original_md = VILANDE_ANALYS_MD.read_text(encoding="utf-8")
        replaced_md = replace_first_example_callout(original_md, callout_lines)
        new_md = replaced_md if replaced_md is not None else build_vilande_analysis_template(callout_lines)
    else:
        original_md = ""
        new_md = build_vilande_analysis_template(callout_lines)

    md_changed = new_md != original_md
    if apply and md_changed:
        VILANDE_ANALYS_MD.write_text(new_md, encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Vilande kursplaner"
    headers = ["Kursplan", "Ämne", "Institution", "Ämnesnamn", "Kursnamn", "Problem", "Länk"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for code, subj_code, inst_code, subject_name, course_name in rows:
        url = KURSPLAN_URL.format(code=code)
        ws.append([
            code,
            subj_code,
            inst_code,
            subject_name,
            course_name,
            "Kursplan finns på du.se men kursen saknas i aktuellt utbud",
            url,
        ])
        row_idx = ws.max_row
        code_cell = ws.cell(row=row_idx, column=1)
        code_cell.hyperlink = url
        code_cell.font = LINK_FONT
        url_cell = ws.cell(row=row_idx, column=7)
        url_cell.hyperlink = url
        url_cell.font = LINK_FONT

    widths = [12, 8, 12, 30, 45, 58, 70]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    if ws.max_row >= 2:
        ws.auto_filter.ref = ws.dimensions

    xlsx_bytes_before = VILANDE_ANALYS_XLSX.read_bytes() if VILANDE_ANALYS_XLSX.exists() else b""
    if apply:
        wb.save(VILANDE_ANALYS_XLSX)
        xlsx_changed = VILANDE_ANALYS_XLSX.read_bytes() != xlsx_bytes_before
    else:
        xlsx_changed = bool(rows) or not VILANDE_ANALYS_XLSX.exists()

    return md_changed, xlsx_changed


# ─────────────────────────────────────────────────────────────────────────────
# Discovery (network)
# ─────────────────────────────────────────────────────────────────────────────

def kursplan_exists(code: str) -> bool:
    """Returnerar True om du.se serverar en riktig kursplansida för koden.

    Söket på du.se listar bara kurser som ges just nu, men en kursplan kan
    fortfarande vara publicerad utan att kursen erbjuds aktuellt. Här
    verifierar vi mot själva kursplanssidan: en giltig sida har ett
    <h1><span property="name">… som söket-fallback inte har.
    """
    soup = fetch_page(SV_URL.format(code=code))
    if soup is None:
        return False
    h1 = soup.find("h1")
    if not h1:
        return False
    return h1.find("span", property="name") is not None


def discover_current_codes(
    only_institutions: set[str] | None,
    only_subjects: set[str] | None,
    quiet: bool,
) -> tuple[dict[str, set[str]], dict[str, dict]]:
    """Returnerar (codes_per_subject, subject_info_per_subject)."""
    codes_per_subject: dict[str, set[str]] = {}
    subject_info: dict[str, dict] = {}

    for inst_code, inst in INSTITUTIONS.items():
        if only_institutions and inst_code not in only_institutions:
            continue
        if not quiet:
            print(f"\n  {CYAN}Institution {inst_code}{RESET}")
        subjects = discover_subjects_for_institution(inst_code, inst)
        for s in subjects:
            subj_code = s.get("code", "")
            if not subj_code:
                continue
            if only_subjects and subj_code not in only_subjects:
                continue
            subject_info[subj_code] = s
            try:
                courses = discover_courses_for_subject(s)
            except Exception as e:
                print(f"  {RED}✗ {subj_code}: {e}{RESET}", file=sys.stderr)
                continue
            codes_per_subject[subj_code] = {c["code"] for c in courses}
            if not quiet:
                print(f"    {subj_code:6s} {s['name']:40s} {len(courses)} aktiva kurser")
            time.sleep(REQUEST_DELAY)
    return codes_per_subject, subject_info


# ─────────────────────────────────────────────────────────────────────────────
# Huvud
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Klassificera kursplaner som aktiv/vilande/ej-aktiv."
    )
    parser.add_argument("--apply", action="store_true",
                        help="Skriv ändringar (annars dry-run).")
    parser.add_argument("--institution", "-i", action="append",
                        help="Begränsa till specifik institution (IIT/IHV/IKS/ISLL).")
    parser.add_argument("--subject", "-s", action="append",
                        help="Begränsa till specifikt ämne (kod, t.ex. DTA).")
    parser.add_argument("--quiet", "-q", action="store_true")
    args = parser.parse_args()

    only_institutions = set(i.upper() for i in args.institution) if args.institution else None
    only_subjects = set(s.upper() for s in args.subject) if args.subject else None

    print(f"{BOLD}{CYAN}Identifiera kursstatus (aktiv/vilande/ej-aktiv){RESET}")
    if not args.apply:
        print(f"  {YELLOW}DRY-RUN — inga filer skrivs{RESET}")
    print(f"  Hämtar nuvarande kurslistor från du.se …")

    codes_per_subject, subject_info = discover_current_codes(
        only_institutions, only_subjects, args.quiet
    )

    print(f"\n{BOLD}Jämför mot vault …{RESET}")
    vault_codes = list_vault_codes_per_subject()

    total_ej_aktiv = 0
    total_vilande = 0
    total_reactivated = 0
    total_moc_writes = 0
    vilande_rows: list[tuple[str, str, str, str, str]] = []

    for subj_code, current_codes in codes_per_subject.items():
        info = subject_info.get(subj_code, {})
        subj_name = info.get("name") or subj_code
        inst_code = info.get("institution", "")
        vault_for_subj = vault_codes.get(subj_code, {})

        suspected = sorted(code for code in vault_for_subj if code not in current_codes)
        reactivated = []
        ej_aktiv_codes: list[str] = []
        vilande_codes: list[str] = []

        # Kurser som åter är i utbudet blir aktiv-status.
        for code in current_codes:
            path = vault_for_subj.get(code)
            if not path:
                continue
            if set_course_status(path, subj_name, "active", args.apply):
                reactivated.append(code)

        # Verifiera kurser som inte är i aktuellt utbud.
        # Finns kursplan på URL: vilande. Annars: ej-aktiv.
        for code in suspected:
            path = vault_for_subj[code]
            if kursplan_exists(code):
                vilande_codes.append(code)
                if set_course_status(path, subj_name, STATUS_VILANDE, args.apply):
                    pass
                vilande_rows.append((code, subj_code, inst_code, subj_name, read_course_name(path)))
            else:
                ej_aktiv_codes.append(code)
                if set_course_status(path, subj_name, STATUS_EJ_AKTIV, args.apply):
                    pass
            time.sleep(REQUEST_DELAY)

        orphan_paths = [vault_for_subj[c] for c in ej_aktiv_codes]

        moc_path = kursplaner_dir(inst_code) / subj_code / f"Ej Aktiv {subj_name} MOC.md"
        if orphan_paths:
            moc_text = build_ej_aktiv_moc(subj_name, subj_code, inst_code, orphan_paths)
            existing = moc_path.read_text(encoding="utf-8") if moc_path.exists() else ""
            if existing != moc_text:
                if args.apply:
                    moc_path.parent.mkdir(parents=True, exist_ok=True)
                    moc_path.write_text(moc_text, encoding="utf-8")
                total_moc_writes += 1
        elif moc_path.exists():
            if args.apply:
                moc_path.unlink()
            total_moc_writes += 1

        if ej_aktiv_codes or vilande_codes or reactivated:
            label = f"{subj_code} ({subj_name})"
            print(f"  {label:50s}  "
                  f"{RED}+{len(ej_aktiv_codes)} ej-aktiv{RESET}  "
                  f"{YELLOW}+{len(vilande_codes)} vilande{RESET}  "
                  f"{GREEN}-{len(reactivated)} reactivated{RESET}")

            for c in ej_aktiv_codes[:5]:
                print(f"      ej-aktiv: {c}")
            if len(ej_aktiv_codes) > 5:
                print(f"      … och {len(ej_aktiv_codes) - 5} till")

            for c in vilande_codes[:5]:
                print(f"      vilande: {c}")
            if len(vilande_codes) > 5:
                print(f"      … och {len(vilande_codes) - 5} till")

        total_ej_aktiv += len(ej_aktiv_codes)
        total_vilande += len(vilande_codes)
        total_reactivated += len(reactivated)

    md_changed, xlsx_changed = write_vilande_analysis(vilande_rows, args.apply)

    print(f"\n{BOLD}Sammanfattning{RESET}")
    print(f"  Ej-aktiva:           {total_ej_aktiv}")
    print(f"  Vilande:             {total_vilande}")
    print(f"  Återaktiverade:      {total_reactivated}")
    print(f"  MOC-skrivningar:     {total_moc_writes}")
    print(f"  Vilande-analys.md:   {'ändrad' if md_changed else 'oförändrad'}")
    print(f"  Vilande-analys.xlsx: {'ändrad' if xlsx_changed else 'oförändrad'}")
    if not args.apply and (total_ej_aktiv or total_vilande or total_reactivated or total_moc_writes or md_changed):
        print(f"\n  Kör igen med {BOLD}--apply{RESET} för att skriva ändringarna.")


if __name__ == "__main__":
    main()
