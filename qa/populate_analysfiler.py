#!/usr/bin/env python3
"""
populate_analysfiler.py — Fyll analysfilerna i
vault-dalarna-university/0X {INST}/Analys/ med fynd från senaste QA-rapporten.

Varje fynd klassas till en institution genom att slå upp kurskoden mot
kursplansfilernas placering i vaulten (med fallback till `institution:` i
frontmatter). Rapportens fynd partitioneras därefter per institution och
skrivs till varje institutions egna analysfiler.

Idempotent: hittar callout-blocket `> [!example]- ... fynd ...` i analysfilen,
och ersätter det med ett nytt block byggt från rapporten. All övrig prosa
(syfte, metod, observationer, rekommendationer) lämnas orörd.

Skriver dessutom en kursspecifik callout överst i varje kursplan som har minst
ett fynd (avgränsad av `<!-- analys:start -->` / `<!-- analys:end -->`), och tar
bort blocket från kursplaner utan kvarvarande fynd. Inga separata analysnoder
skapas — dropdownen ligger direkt i kursplanen. Se ``populate_kursplan_callouts``.

Användning:
    python3 qa/populate_analysfiler.py [--rapport <fil>] [--dry-run]

--rapport   Använd specifik rapport i stället för den senaste.
--dry-run   Visa vad som skulle skrivas, ändra inga filer.
"""
import re
import sys
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT     = Path(__file__).resolve().parent.parent
VAULT         = REPO_ROOT / "vault-dalarna-university"
INST_DIR_NAME = {
    "IIT": "01 IIT",
    "IHV": "02 IHV",
    "IKS": "03 IKS",
    "ISLL": "04 ISLL",
}
RAPPORT_DIR   = Path(__file__).resolve().parent / "rapporter"
RAPPORT_DIR_UTB = Path(__file__).resolve().parent / "rapporter-utb"

COURSE_CODE_RE = re.compile(r"^[A-ZÅÄÖ0-9]{4,8}$")


def institution_analys_dir(inst_code: str) -> Path:
    return VAULT / INST_DIR_NAME[inst_code] / "Analys"


def build_code_to_institution_map() -> dict[str, str]:
    """Skanna vaulten och bygg `kurskod | programkod -> institutionskod`
    baserat på filplaceringen. Både Kursplaner/ och Utbildningsplaner/
    indexeras, så att kursplan-rapportens kursfynd och utbildningsplan-
    rapportens programfynd kan routas till samma institution. Filplacering
    är auktoritativ; vid kollision läses ``institution:`` i frontmatter
    som tiebreaker."""
    mapping: dict[str, str] = {}
    for inst_code, dirname in INST_DIR_NAME.items():
        for sub in ("Kursplaner", "Utbildningsplaner"):
            base = VAULT / dirname / sub
            if not base.exists():
                continue
            for md in base.rglob("*.md"):
                if "MOC" in md.stem:
                    continue
                stem = md.stem
                if not COURSE_CODE_RE.match(stem):
                    continue
                if stem in mapping and mapping[stem] != inst_code:
                    try:
                        text = md.read_text(encoding="utf-8", errors="replace")
                        m = re.search(r'^institution:\s*"?(\w+)"?', text, re.MULTILINE)
                        if m:
                            mapping[stem] = m.group(1)
                    except Exception:
                        pass
                else:
                    mapping[stem] = inst_code
    return mapping

# Analysfil → (rapport-sektionsprefix → "Problem"-etikett i analystabellen)
ANALYS_FILES: dict[str, dict[str, str]] = {
    "Introfras.md": {
        "Introfras före frasning": "Prosa/rubrik före frasning",
    },
    "Frasningskonsistens.md": {
        "Frasning avviker": "Avviker från referensformen",
    },
    "Stavfel och språkbruk.md": {
        "Dubblerat ord":      "Dubblerat ord",
        "Känd felstavning":   "Felstavning",
        "Stavfel (svenska)":  "Felstavning",
        "Stavfel (engelska)": "Felstavning (en)",
    },
    "Betygsskalor.md": {
        "Betygsskala inkonsekvent": "Inkonsekvent delskalor",
    },
    "Examinationsformer.md": {
        "Examinationsformer utan punktlista": "Saknar punktlista",
    },
    "Omfång på lärandemål.md": {
        "För få lärandemål":   "För få mål",
        "För många lärandemål": "För många mål",
        "Långt lärandemål":    "Långt mål",
    },
    "Bloom-taxonomi.md": {
        "Bloom-nivå låg (avancerad kurs)": "Låg verbnivå för avancerad kurs",
        "Bloom-nivå hög (grundkurs)":      "Hög verbnivå för grundkurs",
        "Bloom okänt verb":                "Okänt ledande verb",
    },
    "Samstämmighet svenska och engelska.md": {
        "Paritetsskillnad sv/en": "Paritetsskillnad",
        "Lärandemål saknar punktlista (sv)": "Saknar punktlista (sv)",
        "Lärandemål saknar punktlista (en)": "Saknar punktlista (en)",
    },
    "Betygsrapportering.md": {
        "Betyg saknar punktlista": "Saknar punktlista",
    },
    "Övrigt.md": {
        "Övrigt saknas":           "Sektion saknas",
        "Övrigt utan pedagogiskt-stöd-fras": "Saknar standardfras om pedagogiskt stöd",
    },
    "Förkunskapskrav.md": {
        "Förkunskapskrav saknas":           "Sektion saknas",
        "Förkunskapskrav endast på engelska": "Endast engelsk variant",
        "Förkunskapskrav refererar troligen nedlagd kurs": "Refererar troligen nedlagd kurs",
        "Förkunskapskrav refererar bekräftat nedlagd kurs": "Refererar bekräftat nedlagd kurs",
    },
    "Nedlagda kursreferenser.md": {
        "Nedlagd kursreferens": "Programmet listar nedlagd kurs",
    },
    "Programkurser olänkade.md": {
        "Okänd kursreferens i program":         "Kursnamnet finns varken aktivt eller nedlagt",
        "Aktiv kurs olänkad (scraper-miss)":    "Kurs finns aktivt men scrapern länkade inte",
        "Alternativ-bullet (val mellan kurser)": "Bullet beskriver val mellan flera kurser",
        "Trunkerad kursrad":                    "Kursraden ser avbruten/feltrycklig ut",
        "Programtext skiljer från kursnamn":     "Programtext avviker från kursplanens namn",
    },
}

KURSPLAN_URL = "https://www.du.se/sv/utbildning/kurser/kursplan/?code={code}"
UTBILDNINGSPLAN_URL = "https://www.du.se/sv/utbildning/Program/utbildningsplan/?code={code}"


def build_section_area_map() -> dict[str, tuple[str, str]]:
    """Platta ut ANALYS_FILES till ``rapport-sektion → (område, problem-etikett)``.

    Området är analysfilens namn utan ``.md`` (t.ex. "Stavfel och språkbruk"),
    så att varje rad i en kursplans egen callout pekar tillbaka på motsvarande
    ämnesdropdown. Varje sektion förekommer i exakt en analysfil, så mappningen
    är entydig."""
    m: dict[str, tuple[str, str]] = {}
    for filename, section_map in ANALYS_FILES.items():
        area = filename[:-3] if filename.endswith(".md") else filename
        for section_label, problem_label in section_map.items():
            m[section_label] = (area, problem_label)
    return m


def build_code_to_path_map(subfolder: str) -> dict[str, Path]:
    """Bygg ``kod → planfil`` för alla planer i ``subfolder`` (``Kursplaner`` eller
    ``Utbildningsplaner``) tvärs institutionerna. Hubbar (MOC) och filer vars namn
    inte ser ut som en kod hoppas över. Används för att skriva en plan-specifik
    callout direkt i filen."""
    mapping: dict[str, Path] = {}
    for dirname in INST_DIR_NAME.values():
        base = VAULT / dirname / subfolder
        if not base.exists():
            continue
        for md in base.rglob("*.md"):
            stem = md.stem
            if "MOC" in stem:
                continue
            if not COURSE_CODE_RE.match(stem):
                continue
            mapping[stem] = md
    return mapping


def plan_url_for(subj: str, code: str) -> str:
    """Välj rätt du.se-URL beroende på om raden gäller en kurs- eller
    utbildningsplan. Kontrollerna i ``checks_nedlagda`` sätter ``subj`` till
    ``Utbildningsplaner`` (folder-namnet) för programreferenser."""
    if subj.lower().startswith("utbildning"):
        return UTBILDNINGSPLAN_URL.format(code=code)
    return KURSPLAN_URL.format(code=code)

# Plocka ut det understrukna ordet/uttrycket ur en detaljkolumn — texten inom
# första backtickparet. Används vid deduplicering så att t.ex. en träff från
# "Känd felstavning" och en träff från hunspell-stavning för samma ord i samma
# kursplan räknas som *ett* fynd.
DETAIL_TOKEN_RE = re.compile(r"`([^`]+)`")


def dedup_rows(
    rows: list[tuple[str, str, str, str]],
) -> list[tuple[str, str, str, str]]:
    """En och samma underliggande textuella företeelse räknas bara en gång per
    kursplan. Om samma ord träffas av flera kontroller (t.ex. känd felstavning
    + hunspell) behåller vi det rikare fyndet — det med " → "-rättning vinner."""
    best: dict[tuple[str, str], tuple[str, str, str, str]] = {}
    order: list[tuple[str, str]] = []
    for row in rows:
        code, _subj, _problem, detail = row
        m = DETAIL_TOKEN_RE.search(detail)
        token = m.group(1).lower() if m else detail
        key = (code, token)
        if key not in best:
            best[key] = row
            order.append(key)
        elif "→" in detail and "→" not in best[key][3]:
            best[key] = row
    return [best[k] for k in order]

# ─────────────────────────────────────────────────────────────────────────────
# Rapportparsning
# ─────────────────────────────────────────────────────────────────────────────
SECTION_RE = re.compile(r"^##\s+(.+?)\s*\(\d+")
ROW_RE = re.compile(r"^\|\s*([A-ZÅÄÖ0-9][A-Z0-9]{2,8})\s*\|\s*([\wÅÄÖåäö ]+?)\s*\|(.+?)\|?\s*$")


def parse_rapport(path: Path) -> list[tuple[str, str, str, str]]:
    rows = []
    current_section = "Okänd"
    for line in path.read_text(encoding="utf-8").splitlines():
        m_sec = SECTION_RE.match(line)
        if m_sec:
            current_section = m_sec.group(1).strip()
            continue
        m_row = ROW_RE.match(line)
        if m_row:
            code = m_row.group(1).strip()
            subj = m_row.group(2).strip()
            detail = m_row.group(3).strip().rstrip("|").strip()
            if code.lower() in ("kod", "---"):
                continue
            rows.append((current_section, code, subj, detail))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Callout-byggare
# ─────────────────────────────────────────────────────────────────────────────

DOWNLOAD_ICON_SVG = (
    '<svg class="download-xlsx-icon" width="16" height="16" viewBox="0 0 24 24" '
    'fill="none" stroke="currentColor" stroke-width="2" '
    'stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">'
    '<path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>'
    '<polyline points="7 10 12 15 17 10"/>'
    '<line x1="12" y1="15" x2="12" y2="3"/>'
    '</svg>'
)


def build_callout(
    rows: list[tuple[str, str, str, str]],
    xlsx_filename: str,
    inst_code: str | None = None,
) -> list[str]:
    """rows = [(code, subj, problem_label, detail), ...] -> callout + download link lines.

    With per-institution Analys folders, the same xlsx basename exists in 4 places
    (one per institution). Quartz's "shortest" link resolver sees multiple matches
    and falls back to vault-root → 404. We avoid that by writing the full
    institution-prefixed slug path so transformLink's suffix-match yields a unique hit.
    """
    n = len(rows)
    xlsx_slug = xlsx_filename.replace(" ", "-")  # Quartz slugifies asset names the same way
    if inst_code:
        # e.g. "01 IIT" → "01-IIT" (Quartz slug convention)
        inst_slug = INST_DIR_NAME[inst_code].replace(" ", "-")
        href = f"{inst_slug}/Analys/{xlsx_slug}"
    else:
        href = xlsx_slug
    lines = [
        f'<a class="download-xlsx" href="{href}" download>'
        f'{DOWNLOAD_ICON_SVG}'
        f'<span>Ladda ner som Excel-fil ({n} rader)</span>'
        f'</a>',
        "",
        f"> [!example]- {n} fynd — klicka för att expandera",
        ">",
        "> | Kursplan | Ämne | Problem | Detalj |",
        "> | --- | --- | --- | --- |",
    ]
    for code, subj, problem, detail in rows:
        url = plan_url_for(subj, code)
        # Escape `##` so Quartz doesn't render it as a heading/tag link inside
        # the table cell — the excerpt quotes raw kursplan markdown verbatim.
        cell_detail = detail.replace("##", r"\##")
        lines.append(f"> | [{code}]({url}) | {subj} | {problem} | {cell_detail} |")
    return lines


# ─────────────────────────────────────────────────────────────────────────────
# Kursspecifik callout (skrivs in överst i varje drabbad kursplan)
# ─────────────────────────────────────────────────────────────────────────────
# HTML-kommentarsmarkörer avgränsar blocket så att det kan hittas, ersättas och
# tas bort idempotent. Quartz renderar inte HTML-kommentarer, så de syns aldrig.
KURSPLAN_BLOCK_START = "<!-- analys:start -->"
KURSPLAN_BLOCK_END   = "<!-- analys:end -->"
KURSPLAN_BLOCK_RE = re.compile(
    r"\n*[ \t]*"
    + re.escape(KURSPLAN_BLOCK_START)
    + r".*?"
    + re.escape(KURSPLAN_BLOCK_END)
    + r"[ \t]*\n*",
    re.DOTALL,
)


def strip_kursplan_block(text: str) -> str:
    """Ta bort ett befintligt analysblock (och omgivande blankrader) ur en
    kursplan. Ersätts med en enkel blankradsseparator så att texten är stabil
    vid upprepade körningar."""
    if KURSPLAN_BLOCK_START not in text:
        return text
    return KURSPLAN_BLOCK_RE.sub("\n\n", text)


def build_plan_callout(rows: list[tuple[str, str, str]], noun: str) -> list[str]:
    """rows = [(område, problem, detalj), ...] → markörinramat callout-block.

    Lägger varje kvalitetsnotering för *en* plan i en hopfälld callout. ``noun``
    är "kursplan" eller "utbildningsplan". Området knyter raden till motsvarande
    ämnes-/programanalys-dropdown utan att skapa en separat analysnod i grafen."""
    n = len(rows)
    notering = "kvalitetsnotering" if n == 1 else "kvalitetsnoteringar"
    lines = [
        KURSPLAN_BLOCK_START,
        f"> [!warning]- {n} {notering} i denna {noun} — klicka för att expandera",
        ">",
        "> | Område | Problem | Detalj |",
        "> | --- | --- | --- |",
    ]
    for area, problem, detail in rows:
        # Escapa `##` (rubrik/tagg) och `|` (kolumnavgränsare) så att det citerade
        # kursplansutdraget inte bryter tabellen.
        cell_detail = detail.replace("##", r"\##").replace("|", r"\|")
        lines.append(f"> | {area} | {problem} | {cell_detail} |")
    lines.append(KURSPLAN_BLOCK_END)
    return lines


def insert_kursplan_block(text: str, block_lines: list[str]) -> str:
    """Skriv in ett analysblock direkt efter frontmatter, överst i kursplanen.

    Ett eventuellt tidigare block tas bort först, så funktionen är idempotent."""
    text = strip_kursplan_block(text)
    lines = text.split("\n")
    insert_at = 0
    if lines and lines[0].strip() == "---":
        for i in range(1, len(lines)):
            if lines[i].strip() == "---":
                insert_at = i + 1
                break
    new_lines = lines[:insert_at] + [""] + block_lines + lines[insert_at:]
    return "\n".join(new_lines)


def dedup_plan_rows(
    rows: list[tuple[str, str, str, str]],
) -> list[tuple[str, str, str]]:
    """Deduplicera fynd inom *en* plan, per (område, token) — samma logik som
    analysfilerna använder per fil, så att planens callout speglar unionen av
    analysdropdownarna. rows = [(område, problem, detalj, token), ...]."""
    best: dict[tuple[str, str], tuple[str, str, str]] = {}
    order: list[tuple[str, str]] = []
    for area, problem, detail, token in rows:
        key = (area, token)
        if key not in best:
            best[key] = (area, problem, detail)
            order.append(key)
        elif "→" in detail and "→" not in best[key][2]:
            best[key] = (area, problem, detail)
    return [best[k] for k in order]


def populate_plan_callouts(
    rapport_rows: list[tuple[str, str, str, str]],
    code_to_path: dict[str, Path],
    noun: str,
    dry_run: bool,
) -> None:
    """Skriv/uppdatera/ta bort en plan-specifik callout överst i varje plan i
    ``code_to_path`` som har minst ett fynd. Planer utan fynd får ett eventuellt
    gammalt block borttaget, så funktionen konvergerar mot vault-tillståndet.

    Fynden routas till rätt plan enbart via kodträff mot ``code_to_path`` —
    kurskoder och programkoder lever i skilda mappar, så en kursplan får bara
    kursfynd och en utbildningsplan bara programfynd."""
    section_area = build_section_area_map()

    per_plan: dict[str, list[tuple[str, str, str, str]]] = defaultdict(list)
    for sec, code, subj, detail in rapport_rows:
        area_problem = section_area.get(sec)
        if area_problem is None or code not in code_to_path:
            continue
        area, problem = area_problem
        m = DETAIL_TOKEN_RE.search(detail)
        token = m.group(1).lower() if m else detail
        per_plan[code].append((area, problem, detail, token))

    written = removed = 0
    for code, path in sorted(code_to_path.items()):
        # Läs/skriv som bytes så att radslut (vissa scrapade planer har kvar \r\n
        # i enstaka rader) bevaras exakt — annars skulle read_text normalisera
        # dem och skapa brus utöver själva callout-blocket, och slåss med scrapern.
        original = path.read_bytes().decode("utf-8")
        raw = per_plan.get(code)
        if raw:
            rows = dedup_plan_rows(raw)
            rows.sort(key=lambda r: (r[0], r[1]))
            new_text = insert_kursplan_block(original, build_plan_callout(rows, noun))
        else:
            new_text = strip_kursplan_block(original)

        if new_text == original:
            continue
        if raw:
            written += 1
        else:
            removed += 1
        if not dry_run:
            path.write_bytes(new_text.encode("utf-8"))

    verb = "skulle skrivas" if dry_run else "skrev"
    print(f"\n  {CYAN}{noun.capitalize()}-callouts{RESET}")
    print(f"    {GREEN}{verb} {written}{RESET} {noun}er med fynd, "
          f"{YELLOW}rensade {removed}{RESET} utan kvarvarande fynd")


# ─────────────────────────────────────────────────────────────────────────────
# Excel-export
# ─────────────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="8B1A1A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT   = Font(color="0563C1", underline="single")


def build_xlsx(rows: list[tuple[str, str, str, str]], output_path: Path, sheet_title: str) -> None:
    """Write rows to an .xlsx file with a hyperlinked Kursplan column."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel max sheet title length

    headers = ["Kursplan", "Ämne", "Problem", "Detalj", "Länk"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for code, subj, problem, detail in rows:
        url = plan_url_for(subj, code)
        ws.append([code, subj, problem, detail, url])
        row_idx = ws.max_row
        kod_cell = ws.cell(row=row_idx, column=1)
        kod_cell.hyperlink = url
        kod_cell.font = LINK_FONT
        link_cell = ws.cell(row=row_idx, column=5)
        link_cell.hyperlink = url
        link_cell.font = LINK_FONT

    widths = [12, 8, 28, 70, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    if ws.max_row >= 2:
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# Filersättning
# ─────────────────────────────────────────────────────────────────────────────
CALLOUT_START_RE = re.compile(r"^>\s*\[!example\]")


def replace_callout(text: str, new_block_lines: list[str]) -> str | None:
    """Replace the first `> [!example]…` callout block (and any directly preceding
    .xlsx download link) with new_block_lines.

    Block boundary: starts at the [!example] line by default. If the line(s)
    immediately above (allowing one blank line) contain `.xlsx`, the block start
    is moved up to include them — so the script is idempotent across re-runs.
    Block end: first line after the callout that does not start with `>`.
    """
    lines = text.splitlines()
    start = None
    for i, line in enumerate(lines):
        if CALLOUT_START_RE.match(line):
            start = i
            break
    if start is None:
        return None

    end = start + 1
    while end < len(lines) and lines[end].startswith(">"):
        end += 1

    block_start = start
    j = start - 1
    while j >= 0 and lines[j].strip() == "":
        j -= 1
    if j >= 0 and ".xlsx" in lines[j]:
        block_start = j

    new_lines = lines[:block_start] + new_block_lines + lines[end:]
    return "\n".join(new_lines) + ("\n" if text.endswith("\n") else "")


# ─────────────────────────────────────────────────────────────────────────────
# Färger
# ─────────────────────────────────────────────────────────────────────────────
BOLD   = "\033[1m"
GREEN  = "\033[0;32m"
YELLOW = "\033[0;33m"
CYAN   = "\033[0;36m"
RESET  = "\033[0m"


# ─────────────────────────────────────────────────────────────────────────────
# Huvud
# ─────────────────────────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args

    rapport_path = None
    rapport_path_utb = None
    if "--rapport" in args:
        idx = args.index("--rapport")
        rapport_path = Path(args[idx + 1])
    else:
        # Pick the most recently modified report. Match both timestamped
        # (`rapport-YYYY-MM-DD-HHMM.md`) and bare (`rapport.md`) forms — the
        # latter is the default output of `qa/check_kursplaner.py --out …`
        # so it's often the freshest one.
        rapporter = list(RAPPORT_DIR.glob("rapport*.md"))
        if not rapporter:
            print("Fel: Ingen rapport hittad i qa/rapporter/.", file=sys.stderr)
            sys.exit(1)
        rapport_path = max(rapporter, key=lambda p: p.stat().st_mtime)
        # Same plockning för utbildningsplansrapporten — den är frivillig.
        # Saknas mappen eller är tom hoppar vi bara över utb-fynden.
        if RAPPORT_DIR_UTB.exists():
            rapporter_utb = list(RAPPORT_DIR_UTB.glob("rapport*.md"))
            if rapporter_utb:
                rapport_path_utb = max(rapporter_utb, key=lambda p: p.stat().st_mtime)

    print(f"\n{CYAN}{BOLD}Populera analysfilerna{RESET}")
    print(f"  Rapport (kursplaner): {rapport_path.name}")
    if rapport_path_utb:
        print(f"  Rapport (utb):        {rapport_path_utb.name}")
    if dry_run:
        print(f"  {YELLOW}DRY-RUN — inga filer skrivs{RESET}")
    print()

    rapport_rows = parse_rapport(rapport_path)
    if rapport_path_utb:
        rapport_rows = rapport_rows + parse_rapport(rapport_path_utb)

    by_section: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for sec, code, subj, detail in rapport_rows:
        by_section[sec].append((code, subj, detail))

    code_to_inst = build_code_to_institution_map()
    unmapped: set[str] = set()

    for filename, section_map in ANALYS_FILES.items():
        all_rows: list[tuple[str, str, str, str]] = []
        for section_label, problem_label in section_map.items():
            for code, subj, detail in by_section.get(section_label, []):
                all_rows.append((code, subj, problem_label, detail))

        all_rows = dedup_rows(all_rows)
        all_rows.sort(key=lambda r: (r[1], r[0]))  # by subj, then code

        rows_by_inst: dict[str, list[tuple[str, str, str, str]]] = {
            inst: [] for inst in INST_DIR_NAME
        }
        for row in all_rows:
            code = row[0]
            inst = code_to_inst.get(code)
            if inst is None:
                unmapped.add(code)
                continue
            rows_by_inst.setdefault(inst, []).append(row)

        print(f"  {filename}")
        for inst_code in INST_DIR_NAME:
            analys_path = institution_analys_dir(inst_code) / filename
            rows = rows_by_inst.get(inst_code, [])

            if not analys_path.exists():
                print(f"    {inst_code:<5} {len(rows):>4} fynd  "
                      f"{YELLOW}saknar {filename} — hoppar över{RESET}")
                continue

            xlsx_filename = analys_path.stem + ".xlsx"
            xlsx_path = analys_path.with_suffix(".xlsx")
            callout_lines = build_callout(rows, xlsx_filename, inst_code=inst_code)

            original = analys_path.read_text(encoding="utf-8")
            new_text = replace_callout(original, callout_lines)

            if new_text is None:
                print(f"    {inst_code:<5} {len(rows):>4} fynd  "
                      f"{YELLOW}inget callout-block — hoppar över{RESET}")
                continue

            md_changed = new_text != original
            verb = "skulle skrivas" if dry_run else "skrev"

            if md_changed:
                print(f"    {inst_code:<5} {len(rows):>4} fynd  "
                      f"{GREEN}{verb} md{RESET}")
                if not dry_run:
                    analys_path.write_text(new_text, encoding="utf-8")
            else:
                print(f"    {inst_code:<5} {len(rows):>4} fynd  (md oförändrad)")

            if not dry_run:
                build_xlsx(rows, xlsx_path, sheet_title=analys_path.stem)

    if unmapped:
        print(f"\n  {YELLOW}{len(unmapped)} kurs(er) gick inte att klassa till institution"
              f" och hoppades över: {', '.join(sorted(unmapped))}{RESET}")

    # Skriv den plan-specifika dropdownen överst i varje drabbad kurs- och
    # utbildningsplan. Detta speglar samma fynd som analysfilerna ovan, fast per
    # plan; varje plan får bara de fynd vars kod matchar planen.
    populate_plan_callouts(
        rapport_rows, build_code_to_path_map("Kursplaner"), "kursplan", dry_run
    )
    populate_plan_callouts(
        rapport_rows, build_code_to_path_map("Utbildningsplaner"), "utbildningsplan", dry_run
    )

    print()


if __name__ == "__main__":
    main()
